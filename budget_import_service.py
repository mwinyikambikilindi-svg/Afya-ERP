from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text

import app.extensions as ext


class BudgetImportError(Exception):
    pass


def _new_session():
    if ext.SessionLocal is None:
        raise RuntimeError("Database session factory is not initialized.")
    return ext.SessionLocal()


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _pick_column(header_map: dict[str, int], *names: str) -> int | None:
    for n in names:
        if n in header_map:
            return header_map[n]
    return None


def parse_budget_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise BudgetImportError("Failed to read Excel file. Please upload a valid .xlsx file.") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BudgetImportError("Excel file is empty.")

    header_map = {_normalize_header(v): idx for idx, v in enumerate(rows[0]) if str(v or "").strip()}

    gfs_idx = _pick_column(header_map, "gfs_code", "code", "gfs")
    amt_idx = _pick_column(header_map, "budget_amount", "budget", "amount")
    desc_idx = _pick_column(header_map, "description", "name")

    if gfs_idx is None or amt_idx is None:
        raise BudgetImportError("Excel must contain at least 'GFS Code' and 'Budget Amount' columns.")

    parsed = []
    for excel_row_no, row in enumerate(rows[1:], start=2):
        gfs_code = str(row[gfs_idx] or "").strip()
        description = str(row[desc_idx] or "").strip() if desc_idx is not None else ""
        raw_amount = row[amt_idx]
        if not gfs_code and (raw_amount is None or str(raw_amount).strip() == ""):
            continue
        parsed.append({
            "excel_row_no": excel_row_no,
            "gfs_code": gfs_code,
            "description": description,
            "raw_amount": raw_amount,
        })
    return parsed


def import_budget_excel(budget_header_id: int, file_bytes: bytes) -> dict[str, Any]:
    rows = parse_budget_excel(file_bytes)
    session = _new_session()

    processed = 0
    imported = 0
    skipped = 0
    errors: list[str] = []

    try:
        gfs_rows = session.execute(text("SELECT id, code, name FROM gfs_codes")).mappings().all()
        gfs_index = {str(r["code"]).strip().upper(): dict(r) for r in gfs_rows}

        header = session.execute(text("SELECT id FROM budget_headers WHERE id = :budget_header_id"), {"budget_header_id": budget_header_id}).first()
        if not header:
            raise BudgetImportError("Selected budget header was not found.")

        for row in rows:
            processed += 1
            gfs_code = str(row["gfs_code"]).strip().upper()
            if not gfs_code:
                skipped += 1
                errors.append(f"Row {row['excel_row_no']}: Missing GFS Code.")
                continue

            gfs = gfs_index.get(gfs_code)
            if not gfs:
                skipped += 1
                errors.append(f"Row {row['excel_row_no']}: GFS Code '{gfs_code}' not found.")
                continue

            raw_amount = row["raw_amount"]
            try:
                amount = Decimal(str(raw_amount).replace(",", "").strip())
            except (InvalidOperation, AttributeError):
                skipped += 1
                errors.append(f"Row {row['excel_row_no']}: Invalid Budget Amount '{raw_amount}'.")
                continue

            session.execute(text("""
                INSERT INTO budget_lines (budget_header_id, gfs_code_id, budget_amount)
                VALUES (:budget_header_id, :gfs_code_id, :budget_amount)
                ON CONFLICT (budget_header_id, gfs_code_id)
                DO UPDATE SET budget_amount = EXCLUDED.budget_amount
            """), {
                "budget_header_id": budget_header_id,
                "gfs_code_id": gfs["id"],
                "budget_amount": amount,
            })
            imported += 1

        session.commit()
        return {"processed": processed, "imported": imported, "skipped": skipped, "errors": errors}
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
