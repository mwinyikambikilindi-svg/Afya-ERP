from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from sqlalchemy import text
import app.extensions as ext


class BudgetReportingError(Exception):
    pass


def _new_session():
    if ext.SessionLocal is None:
        raise RuntimeError("Database session factory is not initialized.")
    return ext.SessionLocal()


def _get_columns(session, table_name: str) -> set[str]:
    rows = session.execute(
        text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = :table_name
            ORDER BY ordinal_position
        """),
        {"table_name": table_name},
    ).fetchall()
    return {r[0] for r in rows}


def _pick_first_existing(columns: set[str], *candidates: str) -> str:
    for c in candidates:
        if c in columns:
            return c
    return ""


def _load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        return Workbook, Font
    except Exception as exc:
        raise BudgetReportingError("Excel export requires openpyxl. Install it with: pip install openpyxl") from exc


SECTION_ORDER = ["Revenue","Expenses","Current Assets","Non-Current Assets","Current Liabilities","Net Assets/Equity","Unclassified"]


def list_fiscal_years():
    session = _new_session()
    try:
        fy_cols = _get_columns(session, "fiscal_years")
        if not fy_cols:
            return []
        name_col = _pick_first_existing(fy_cols, "name", "year_name")
        start_col = _pick_first_existing(fy_cols, "start_date")
        end_col = _pick_first_existing(fy_cols, "end_date")
        active_col = _pick_first_existing(fy_cols, "is_active")
        select_parts = ["id"]
        select_parts.append(f"{name_col} AS name" if name_col else "id::text AS name")
        select_parts.append(f"{start_col} AS start_date" if start_col else "NULL::date AS start_date")
        select_parts.append(f"{end_col} AS end_date" if end_col else "NULL::date AS end_date")
        select_parts.append(f"COALESCE({active_col}, false) AS is_active" if active_col else "false AS is_active")
        sql = f"SELECT {', '.join(select_parts)} FROM fiscal_years ORDER BY id DESC"
        rows = session.execute(text(sql)).mappings().all()
        return [dict(r) for r in rows]
    finally:
        session.close()


def get_budget_headers(fiscal_year_id: int | None = None):
    session = _new_session()
    try:
        filters = []
        params = {}
        if fiscal_year_id:
            filters.append("fiscal_year_id = :fiscal_year_id")
            params["fiscal_year_id"] = fiscal_year_id
        where_clause = ("WHERE " + " AND ".join(filters)) if filters else ""
        sql = f"""
            SELECT id, fiscal_year_id, budget_name, status, created_at, approved_at
            FROM budget_headers
            {where_clause}
            ORDER BY
                CASE WHEN status = 'approved' THEN 0 ELSE 1 END,
                approved_at DESC NULLS LAST,
                id DESC
        """
        rows = session.execute(text(sql), params).mappings().all()
        return [dict(r) for r in rows]
    finally:
        session.close()


def _journal_context(session):
    jl_cols = _get_columns(session, "journal_lines")
    ga_cols = _get_columns(session, "gl_accounts")
    jb_cols = _get_columns(session, "journal_batches")
    map_cols = _get_columns(session, "gl_account_gfs_map")
    gfs_cols = _get_columns(session, "gfs_codes")
    ap_cols = _get_columns(session, "accounting_periods")
    return {
        "jl_gl_fk": _pick_first_existing(jl_cols, "gl_account_id", "account_id"),
        "jl_debit": _pick_first_existing(jl_cols, "debit", "debit_amount"),
        "jl_credit": _pick_first_existing(jl_cols, "credit", "credit_amount"),
        "jl_batch_fk": _pick_first_existing(jl_cols, "journal_batch_id", "batch_id"),
        "ga_id": _pick_first_existing(ga_cols, "id"),
        "jb_status": _pick_first_existing(jb_cols, "status"),
        "jb_fy": _pick_first_existing(jb_cols, "fiscal_year_id"),
        "jb_period_fk": _pick_first_existing(jb_cols, "accounting_period_id", "period_id"),
        "gm_ga_fk": _pick_first_existing(map_cols, "gl_account_id"),
        "gm_gfs_fk": _pick_first_existing(map_cols, "gfs_code_id"),
        "g_id": _pick_first_existing(gfs_cols, "id"),
        "g_code": _pick_first_existing(gfs_cols, "code"),
        "g_name": _pick_first_existing(gfs_cols, "name"),
        "g_type": _pick_first_existing(gfs_cols, "gfs_type"),
        "g_section": _pick_first_existing(gfs_cols, "statement_section"),
        "ap_id": _pick_first_existing(ap_cols, "id"),
        "ap_fy": _pick_first_existing(ap_cols, "fiscal_year_id"),
    }


def _build_fiscal_year_filter(ctx: dict) -> tuple[str, str]:
    if ctx["jb_fy"]:
        return "", f"jb.{ctx['jb_fy']} = :fiscal_year_id"
    if ctx["jb_period_fk"] and ctx["ap_id"] and ctx["ap_fy"]:
        return f"JOIN accounting_periods ap ON ap.{ctx['ap_id']} = jb.{ctx['jb_period_fk']}", f"ap.{ctx['ap_fy']} = :fiscal_year_id"
    raise BudgetReportingError("Could not determine fiscal-year linkage for journal_batches. Expected either journal_batches.fiscal_year_id or a link to accounting_periods.")


def get_budget_vs_actual_statement(fiscal_year_id: int, budget_header_id: int | None = None):
    if not fiscal_year_id:
        raise BudgetReportingError("Fiscal year is required for Budget vs Actual report.")
    session = _new_session()
    try:
        ctx = _journal_context(session)
        required = [ctx["jl_gl_fk"],ctx["jl_debit"],ctx["jl_credit"],ctx["jl_batch_fk"],ctx["ga_id"],ctx["gm_ga_fk"],ctx["gm_gfs_fk"],ctx["g_id"],ctx["g_code"],ctx["g_name"],ctx["g_type"],ctx["g_section"],ctx["jb_status"]]
        if not all(required):
            raise BudgetReportingError("Schema is missing expected columns for Budget vs Actual report.")
        actual_join_sql, actual_fy_condition = _build_fiscal_year_filter(ctx)

        if budget_header_id is None:
            hdr = session.execute(text("""
                SELECT id
                FROM budget_headers
                WHERE fiscal_year_id = :fiscal_year_id
                  AND status = 'approved'
                ORDER BY approved_at DESC NULLS LAST, id DESC
                LIMIT 1
            """), {"fiscal_year_id": fiscal_year_id}).first()
            if not hdr:
                hdr = session.execute(text("""
                    SELECT id
                    FROM budget_headers
                    WHERE fiscal_year_id = :fiscal_year_id
                    ORDER BY id DESC
                    LIMIT 1
                """), {"fiscal_year_id": fiscal_year_id}).first()
            if hdr:
                budget_header_id = hdr[0]

        budget_rows = []
        if budget_header_id:
            budget_rows = session.execute(text(f"""
                SELECT
                    g.{ctx['g_code']} AS gfs_code,
                    g.{ctx['g_name']} AS gfs_name,
                    COALESCE(g.{ctx['g_section']}, 'Unclassified') AS statement_section,
                    g.{ctx['g_type']} AS gfs_type,
                    COALESCE(bl.budget_amount, 0) AS budget_amount
                FROM budget_lines bl
                JOIN gfs_codes g ON g.{ctx['g_id']} = bl.gfs_code_id
                WHERE bl.budget_header_id = :budget_header_id
                ORDER BY COALESCE(g.{ctx['g_section']}, 'Unclassified'), g.{ctx['g_code']}
            """), {"budget_header_id": budget_header_id}).mappings().all()

        actual_rows = session.execute(text(f"""
            SELECT
                g.{ctx['g_code']} AS gfs_code,
                g.{ctx['g_name']} AS gfs_name,
                COALESCE(g.{ctx['g_section']}, 'Unclassified') AS statement_section,
                g.{ctx['g_type']} AS gfs_type,
                CASE
                    WHEN g.{ctx['g_type']} IN ('asset', 'expenditure') THEN COALESCE(SUM(jl.{ctx['jl_debit']}) - SUM(jl.{ctx['jl_credit']}), 0)
                    WHEN g.{ctx['g_type']} IN ('liability', 'equity', 'revenue', 'contra_asset') THEN COALESCE(SUM(jl.{ctx['jl_credit']}) - SUM(jl.{ctx['jl_debit']}), 0)
                    ELSE COALESCE(SUM(jl.{ctx['jl_debit']}) - SUM(jl.{ctx['jl_credit']}), 0)
                END AS actual_amount
            FROM journal_lines jl
            JOIN journal_batches jb ON jb.id = jl.{ctx['jl_batch_fk']}
            {actual_join_sql}
            JOIN gl_accounts ga ON ga.{ctx['ga_id']} = jl.{ctx['jl_gl_fk']}
            JOIN gl_account_gfs_map gm ON gm.{ctx['gm_ga_fk']} = ga.{ctx['ga_id']}
            JOIN gfs_codes g ON g.{ctx['g_id']} = gm.{ctx['gm_gfs_fk']}
            WHERE COALESCE(jb.{ctx['jb_status']}, 'posted') NOT IN ('draft', 'cancelled')
              AND {actual_fy_condition}
            GROUP BY g.{ctx['g_code']}, g.{ctx['g_name']}, g.{ctx['g_type']}, g.{ctx['g_section']}
            ORDER BY COALESCE(g.{ctx['g_section']}, 'Unclassified'), g.{ctx['g_code']}
        """), {"fiscal_year_id": fiscal_year_id}).mappings().all()

        budget_index = {r["gfs_code"]: Decimal(str(r.get("budget_amount", 0) or 0)) for r in budget_rows}
        actual_index = {r["gfs_code"]: dict(r) for r in actual_rows}
        all_codes = sorted(set(budget_index.keys()) | set(actual_index.keys()))
        rows = []
        for code in all_codes:
            actual = actual_index.get(code)
            budget_amt = budget_index.get(code, Decimal("0.00"))
            actual_amt = Decimal(str((actual or {}).get("actual_amount", 0) or 0))
            variance_amt = actual_amt - budget_amt
            variance_pct = (variance_amt / budget_amt * Decimal("100")) if budget_amt != 0 else Decimal("0.00")
            rows.append({
                "gfs_code": code,
                "gfs_name": (actual or {}).get("gfs_name") or next((r["gfs_name"] for r in budget_rows if r["gfs_code"] == code), ""),
                "statement_section": (actual or {}).get("statement_section") or next((r["statement_section"] for r in budget_rows if r["gfs_code"] == code), "Unclassified"),
                "budget_amount": budget_amt,
                "actual_amount": actual_amt,
                "variance_amount": variance_amt,
                "variance_pct": variance_pct,
                "variance_flag": "positive" if variance_amt > 0 else "negative" if variance_amt < 0 else "neutral",
            })

        sections = {}
        for r in rows:
            sec = r["statement_section"] or "Unclassified"
            sections.setdefault(sec, {"section": sec, "rows": [], "budget_total": Decimal("0.00"), "actual_total": Decimal("0.00"), "variance_total": Decimal("0.00")})
            sections[sec]["rows"].append(r)
            sections[sec]["budget_total"] += Decimal(str(r["budget_amount"]))
            sections[sec]["actual_total"] += Decimal(str(r["actual_amount"]))
            sections[sec]["variance_total"] += Decimal(str(r["variance_amount"]))

        ordered = sorted(list(sections.values()), key=lambda x: (SECTION_ORDER.index(x["section"]) if x["section"] in SECTION_ORDER else 999, x["section"]))
        return {
            "budget_header_id": budget_header_id,
            "sections": ordered,
            "grand_budget": sum((s["budget_total"] for s in ordered), Decimal("0.00")),
            "grand_actual": sum((s["actual_total"] for s in ordered), Decimal("0.00")),
            "grand_variance": sum((s["variance_total"] for s in ordered), Decimal("0.00")),
        }
    finally:
        session.close()


def export_budget_vs_actual_to_excel(title: str, result: dict):
    Workbook, Font = _load_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget vs Actual"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    row_no = 3
    headers = ["Code", "Description", "Budget", "Actual", "Variance", "Variance %"]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=row_no, column=idx, value=h).font = Font(bold=True)
    row_no += 1
    for section in result["sections"]:
        ws.cell(row=row_no, column=1, value=section["section"]).font = Font(bold=True)
        row_no += 1
        for item in section["rows"]:
            ws.cell(row=row_no, column=1, value=item.get("gfs_code"))
            ws.cell(row=row_no, column=2, value=item.get("gfs_name"))
            ws.cell(row=row_no, column=3, value=float(item.get("budget_amount", 0) or 0))
            ws.cell(row=row_no, column=4, value=float(item.get("actual_amount", 0) or 0))
            ws.cell(row=row_no, column=5, value=float(item.get("variance_amount", 0) or 0))
            ws.cell(row=row_no, column=6, value=float(item.get("variance_pct", 0) or 0))
            row_no += 1
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
