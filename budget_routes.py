from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from flask import (
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

import app.extensions as ext
from app.services.auth_service import login_required, role_required
from app.services.budget_import_service import import_budget_excel
from app.services.budget_workflow_service import (
    BudgetWorkflowError,
    approve_budget,
    assert_budget_editable,
    return_budget_to_draft,
    submit_budget,
)


def _new_session():
    if ext.SessionLocal is None:
        raise RuntimeError("Database session factory is not initialized.")
    return ext.SessionLocal()


def _get_columns(session, table_name: str) -> set[str]:
    rows = session.execute(
        text(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = :table_name
            ORDER BY ordinal_position
            """
        ),
        {"table_name": table_name},
    ).fetchall()
    return {r[0] for r in rows}


def _pick_first_existing(columns: set[str], *candidates: str) -> str:
    for c in candidates:
        if c in columns:
            return c
    return ""


def _fy_name_expr(session, alias: str = "fy") -> str:
    fy_cols = _get_columns(session, "fiscal_years")
    fy_name_col = _pick_first_existing(fy_cols, "name", "year_name")
    if fy_name_col:
        return f"{alias}.{fy_name_col}"
    return f"CAST({alias}.id AS VARCHAR)"


def _list_fiscal_years(session):
    fy_cols = _get_columns(session, "fiscal_years")
    name_col = _pick_first_existing(fy_cols, "name", "year_name")

    if name_col:
        sql = f"""
            SELECT id, {name_col} AS name
            FROM fiscal_years
            ORDER BY id DESC
        """
    else:
        sql = """
            SELECT id, CAST(id AS VARCHAR) AS name
            FROM fiscal_years
            ORDER BY id DESC
        """

    rows = session.execute(text(sql)).mappings().all()
    return [dict(r) for r in rows]


def _list_budget_headers(session):
    fy_name_expr = _fy_name_expr(session, "fy")
    rows = session.execute(
        text(
            f"""
            SELECT
                bh.id,
                bh.fiscal_year_id,
                {fy_name_expr} AS fiscal_year_name,
                bh.budget_name,
                bh.status,
                bh.created_at,
                bh.submitted_at,
                bh.approved_at
            FROM budget_headers bh
            JOIN fiscal_years fy ON fy.id = bh.fiscal_year_id
            ORDER BY bh.id DESC
            """
        )
    ).mappings().all()
    return [dict(r) for r in rows]


def _list_gfs_codes(session):
    rows = session.execute(
        text(
            """
            SELECT
                id,
                code,
                name,
                COALESCE(statement_section, '') AS statement_section
            FROM gfs_codes
            WHERE COALESCE(is_active, true) = true
            ORDER BY COALESCE(statement_section, ''), code
            """
        )
    ).mappings().all()
    return [dict(r) for r in rows]


def _get_budget_header(session, header_id: int):
    fy_name_expr = _fy_name_expr(session, "fy")
    row = session.execute(
        text(
            f"""
            SELECT
                bh.id,
                bh.fiscal_year_id,
                {fy_name_expr} AS fiscal_year_name,
                bh.budget_name,
                bh.status,
                bh.created_at,
                bh.submitted_at,
                bh.approved_at
            FROM budget_headers bh
            JOIN fiscal_years fy ON fy.id = bh.fiscal_year_id
            WHERE bh.id = :header_id
            """
        ),
        {"header_id": header_id},
    ).mappings().first()
    return dict(row) if row else None


def _list_budget_lines(session, header_id: int):
    rows = session.execute(
        text(
            """
            SELECT
                bl.id,
                bl.budget_header_id,
                bl.gfs_code_id,
                g.code AS gfs_code,
                g.name AS gfs_name,
                COALESCE(g.statement_section, '') AS statement_section,
                bl.budget_amount
            FROM budget_lines bl
            JOIN gfs_codes g ON g.id = bl.gfs_code_id
            WHERE bl.budget_header_id = :header_id
            ORDER BY COALESCE(g.statement_section, ''), g.code
            """
        ),
        {"header_id": header_id},
    ).mappings().all()
    return [dict(r) for r in rows]


def _build_budget_master_template(session) -> BytesIO:
    rows = session.execute(
        text(
            """
            SELECT
                code,
                name,
                COALESCE(statement_section, '') AS statement_section
            FROM gfs_codes
            WHERE COALESCE(is_active, true) = true
            ORDER BY COALESCE(statement_section, ''), code
            """
        )
    ).mappings().all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Budget_Input"

    headers = [
        "GFS Code",
        "GFS Name",
        "Statement Section",
        "Budget Amount (TZS)",
        "Notes (optional)",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0B2E4F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [16, 42, 24, 20, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(
            [
                row["code"],
                row["name"],
                row["statement_section"],
                None,
                "",
            ]
        )

    for row_no in range(2, ws.max_row + 1):
        ws.cell(row=row_no, column=4).number_format = "#,##0.00"

    ref = wb.create_sheet("GFS_Reference")
    ref_headers = ["GFS Code", "GFS Name", "Statement Section"]
    ref.append(ref_headers)

    for col, heading in enumerate(ref_headers, start=1):
        cell = ref.cell(row=1, column=col, value=heading)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ref.column_dimensions["A"].width = 16
    ref.column_dimensions["B"].width = 48
    ref.column_dimensions["C"].width = 24
    ref.freeze_panes = "A2"

    for row in rows:
        ref.append(
            [
                row["code"],
                row["name"],
                row["statement_section"],
            ]
        )

    ins = wb.create_sheet("Instructions")
    ins["A1"] = "AFYA ERP - Dynamic GFS Budget Master Template"
    ins["A1"].font = Font(bold=True, size=14)
    ins["A3"] = "1) Do not change the first row headers in Budget_Input."
    ins["A4"] = "2) Enter budget amounts only in 'Budget Amount (TZS)'."
    ins["A5"] = "3) You may add notes in the Notes column if needed."
    ins["A6"] = "4) Keep the GFS Code values exactly as generated."
    ins["A7"] = "5) Save as .xlsx and upload through Budget Import."
    ins["A9"] = "This template is generated from the live GFS codes in AFYA ERP."
    ins.column_dimensions["A"].width = 100

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def register_budget_routes(flask_app):
    @flask_app.route("/budget")
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_index():
        session = _new_session()
        try:
            headers = _list_budget_headers(session)
            return render_template("budget/index.html", headers=headers)
        finally:
            session.close()

    @flask_app.route("/budget/create", methods=["GET", "POST"])
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_create():
        session = _new_session()
        try:
            if request.method == "POST":
                fiscal_year_id = request.form.get("fiscal_year_id", type=int)
                budget_name = (request.form.get("budget_name") or "").strip()
                status = (request.form.get("status") or "draft").strip()

                if not fiscal_year_id:
                    flash("Fiscal year is required.", "error")
                    return redirect(url_for("budget_create"))

                if not budget_name:
                    flash("Budget name is required.", "error")
                    return redirect(url_for("budget_create"))

                session.execute(
                    text(
                        """
                        INSERT INTO budget_headers (fiscal_year_id, budget_name, status)
                        VALUES (:fiscal_year_id, :budget_name, :status)
                        """
                    ),
                    {
                        "fiscal_year_id": fiscal_year_id,
                        "budget_name": budget_name,
                        "status": status,
                    },
                )
                session.commit()
                flash("Budget header created successfully.", "success")
                return redirect(url_for("budget_index"))

            fiscal_years = _list_fiscal_years(session)
            return render_template("budget/create.html", fiscal_years=fiscal_years)
        except Exception:
            session.rollback()
            current_app.logger.exception("Budget create error")
            flash("Failed to create budget header.", "error")
            return redirect(url_for("budget_index"))
        finally:
            session.close()

    @flask_app.route("/budget/<int:header_id>/lines", methods=["GET", "POST"])
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_lines(header_id: int):
        session = _new_session()
        try:
            header = _get_budget_header(session, header_id)
            if not header:
                flash("Budget header not found.", "error")
                return redirect(url_for("budget_index"))

            if request.method == "POST":
                gfs_code_id = request.form.get("gfs_code_id", type=int)
                amount_raw = (request.form.get("budget_amount") or "0").replace(",", "").strip()

                assert_budget_editable(session, header_id)

                if not gfs_code_id:
                    flash("GFS code is required.", "error")
                    return redirect(url_for("budget_lines", header_id=header_id))

                try:
                    amount = Decimal(amount_raw or "0")
                except InvalidOperation:
                    flash("Budget amount is invalid.", "error")
                    return redirect(url_for("budget_lines", header_id=header_id))

                session.execute(
                    text(
                        """
                        INSERT INTO budget_lines (budget_header_id, gfs_code_id, budget_amount)
                        VALUES (:budget_header_id, :gfs_code_id, :budget_amount)
                        ON CONFLICT (budget_header_id, gfs_code_id)
                        DO UPDATE SET budget_amount = EXCLUDED.budget_amount
                        """
                    ),
                    {
                        "budget_header_id": header_id,
                        "gfs_code_id": gfs_code_id,
                        "budget_amount": amount,
                    },
                )
                session.commit()
                flash("Budget line saved successfully.", "success")
                return redirect(url_for("budget_lines", header_id=header_id))

            lines = _list_budget_lines(session, header_id)
            gfs_codes = _list_gfs_codes(session)
            total_budget = sum(Decimal(str(r["budget_amount"] or 0)) for r in lines)

            return render_template(
                "budget/lines.html",
                header=header,
                lines=lines,
                gfs_codes=gfs_codes,
                total_budget=total_budget,
            )
        except BudgetWorkflowError as e:
            session.rollback()
            flash(str(e), "error")
            return redirect(url_for("budget_lines", header_id=header_id))
        except Exception:
            session.rollback()
            current_app.logger.exception("Budget lines error")
            flash("Failed to load or save budget lines.", "error")
            return redirect(url_for("budget_index"))
        finally:
            session.close()

    @flask_app.route("/budget/line/<int:line_id>/delete", methods=["POST"])
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_line_delete(line_id: int):
        session = _new_session()
        try:
            row = session.execute(
                text(
                    """
                    SELECT budget_header_id
                    FROM budget_lines
                    WHERE id = :line_id
                    """
                ),
                {"line_id": line_id},
            ).first()

            if not row:
                flash("Budget line not found.", "error")
                return redirect(url_for("budget_index"))

            header_id = row[0]
            assert_budget_editable(session, header_id)

            session.execute(
                text(
                    """
                    DELETE FROM budget_lines
                    WHERE id = :line_id
                    """
                ),
                {"line_id": line_id},
            )
            session.commit()

            flash("Budget line deleted successfully.", "success")
            return redirect(url_for("budget_lines", header_id=header_id))
        except BudgetWorkflowError as e:
            session.rollback()
            flash(str(e), "error")
            return redirect(url_for("budget_index"))
        except Exception:
            session.rollback()
            current_app.logger.exception("Budget line delete error")
            flash("Failed to delete budget line.", "error")
            return redirect(url_for("budget_index"))
        finally:
            session.close()

    @flask_app.route("/budget/import", methods=["GET", "POST"])
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_import():
        session = _new_session()
        result = None
        try:
            headers = _list_budget_headers(session)

            if request.method == "POST":
                budget_header_id = request.form.get("budget_header_id", type=int)
                upload = request.files.get("budget_file")

                if not budget_header_id:
                    flash("Budget header is required.", "error")
                    return redirect(url_for("budget_import"))

                if not upload or not upload.filename:
                    flash("Please choose an Excel file.", "error")
                    return redirect(url_for("budget_import"))

                assert_budget_editable(session, budget_header_id)

                result = import_budget_excel(
                    budget_header_id=budget_header_id,
                    file_bytes=upload.read(),
                )
                flash("Budget Excel imported successfully.", "success")

            return render_template("budget/import.html", headers=headers, result=result)
        except BudgetWorkflowError as e:
            flash(str(e), "error")
            return redirect(url_for("budget_import"))
        except Exception:
            current_app.logger.exception("Budget import error")
            flash("Failed to import budget Excel.", "error")
            return redirect(url_for("budget_import"))
        finally:
            session.close()

    @flask_app.route("/budget/template/download")
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_template_download():
        template_dir = Path(flask_app.root_path) / "static" / "templates"
        filename = "Budget_Master_Template_GFS_Assisted_AFYA_ERP_FINAL.xlsx"

        return send_from_directory(
            directory=str(template_dir),
            path=filename,
            as_attachment=True,
            download_name=filename,
        )

    @flask_app.route("/budget/template/dynamic-download")
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_template_dynamic_download():
        session = _new_session()
        try:
            stream = _build_budget_master_template(session)
            return send_file(
                stream,
                as_attachment=True,
                download_name="Budget_Master_Template_GFS_Live_AFYA_ERP.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        finally:
            session.close()

    @flask_app.route("/budget/<int:header_id>/submit", methods=["POST"])
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_submit(header_id: int):
        try:
            submit_budget(header_id)
            flash("Budget submitted successfully.", "success")
        except BudgetWorkflowError as e:
            flash(str(e), "error")
        except Exception:
            current_app.logger.exception("Budget submit error")
            flash("Failed to submit budget.", "error")
        return redirect(url_for("budget_lines", header_id=header_id))

    @flask_app.route("/budget/<int:header_id>/return-to-draft", methods=["POST"])
    @login_required
    @role_required("ADMIN", "ACCOUNTANT", "MANAGER")
    def budget_return_to_draft(header_id: int):
        try:
            return_budget_to_draft(header_id)
            flash("Budget returned to draft.", "success")
        except BudgetWorkflowError as e:
            flash(str(e), "error")
        except Exception:
            current_app.logger.exception("Budget return-to-draft error")
            flash("Failed to return budget to draft.", "error")
        return redirect(url_for("budget_lines", header_id=header_id))

    @flask_app.route("/budget/<int:header_id>/approve", methods=["POST"])
    @login_required
    @role_required("ADMIN", "MANAGER")
    def budget_approve(header_id: int):
        try:
            approve_budget(header_id)
            flash("Budget approved successfully.", "success")
        except BudgetWorkflowError as e:
            flash(str(e), "error")
        except Exception:
            current_app.logger.exception("Budget approve error")
            flash("Failed to approve budget.", "error")
        return redirect(url_for("budget_lines", header_id=header_id))