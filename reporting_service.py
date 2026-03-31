from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any
from sqlalchemy import text
def _load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        return Workbook, Font
    except Exception as exc:
        raise ReportingServiceError(
            "Excel export requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

import app.extensions as ext


class ReportingServiceError(Exception):
    pass


def _new_session():
    if ext.SessionLocal is None:
        raise RuntimeError("Database session factory is not initialized.")
    return ext.SessionLocal()


def _get_columns(session, table_name: str) -> set[str]:
    rows = session.execute(
        text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = :table_name
            ORDER BY ordinal_position
        """),
        {"table_name": table_name},
    ).fetchall()
    return {r[0] for r in rows}


def _pick_first_existing(columns: set[str], *candidates: str) -> str:
    for c in candidates:
        if c in columns:
            return c
    return ""


SECTION_ORDER = {
    "SOFP": ["Current Assets", "Non-Current Assets", "Current Liabilities", "Net Assets/Equity", "Unclassified"],
    "SOFPERF": ["Revenue", "Expenses", "Unclassified"],
}


def list_fiscal_years():
    session = _new_session()
    try:
        fy_cols = _get_columns(session, "fiscal_years")
        if not fy_cols:
            return []
        name_col = _pick_first_existing(fy_cols, "name", "year_name")
        start_col = _pick_first_existing(fy_cols, "start_date")
        end_col = _pick_first_existing(fy_cols, "end_date")
        active_col = _pick_first_existing(fy_cols, "is_active")
        select_parts = ["id"]
        select_parts.append(f"{name_col} AS name" if name_col else "id::text AS name")
        select_parts.append(f"{start_col} AS start_date" if start_col else "NULL::date AS start_date")
        select_parts.append(f"{end_col} AS end_date" if end_col else "NULL::date AS end_date")
        select_parts.append(f"COALESCE({active_col}, false) AS is_active" if active_col else "false AS is_active")
        sql = f"SELECT {', '.join(select_parts)} FROM fiscal_years ORDER BY id DESC"
        rows = session.execute(text(sql)).mappings().all()
        return [dict(r) for r in rows]
    finally:
        session.close()


def get_prior_fiscal_year_id(current_fiscal_year_id: int | None):
    if current_fiscal_year_id is None:
        return None
    rows = list_fiscal_years()
    ids = [r["id"] for r in rows]
    if current_fiscal_year_id not in ids:
        return None
    idx = ids.index(current_fiscal_year_id)
    return ids[idx + 1] if idx + 1 < len(ids) else None


def _journal_context(session):
    jl_cols = _get_columns(session, "journal_lines")
    ga_cols = _get_columns(session, "gl_accounts")
    jb_cols = _get_columns(session, "journal_batches")
    map_cols = _get_columns(session, "gl_account_gfs_map")
    gfs_cols = _get_columns(session, "gfs_codes")
    return {
        "jl_gl_fk": _pick_first_existing(jl_cols, "gl_account_id", "account_id"),
        "jl_debit": _pick_first_existing(jl_cols, "debit", "debit_amount"),
        "jl_credit": _pick_first_existing(jl_cols, "credit", "credit_amount"),
        "jl_batch_fk": _pick_first_existing(jl_cols, "journal_batch_id", "batch_id"),
        "ga_id": _pick_first_existing(ga_cols, "id"),
        "ga_code": _pick_first_existing(ga_cols, "account_code", "code", "account_no", "account_number"),
        "ga_name": _pick_first_existing(ga_cols, "account_name", "name", "description"),
        "jb_status": _pick_first_existing(jb_cols, "status"),
        "jb_fy": _pick_first_existing(jb_cols, "fiscal_year_id"),
        "jb_date": _pick_first_existing(jb_cols, "posting_date", "journal_date", "transaction_date", "entry_date"),
        "gm_ga_fk": _pick_first_existing(map_cols, "gl_account_id"),
        "gm_gfs_fk": _pick_first_existing(map_cols, "gfs_code_id"),
        "g_id": _pick_first_existing(gfs_cols, "id"),
        "g_code": _pick_first_existing(gfs_cols, "code"),
        "g_name": _pick_first_existing(gfs_cols, "name"),
        "g_type": _pick_first_existing(gfs_cols, "gfs_type"),
        "g_stmt": _pick_first_existing(gfs_cols, "statement_name"),
        "g_section": _pick_first_existing(gfs_cols, "statement_section"),
        "g_note": _pick_first_existing(gfs_cols, "note_no"),
        "g_group": _pick_first_existing(gfs_cols, "report_group"),
    }


def _sort_sections(statement_name: str, sections: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = SECTION_ORDER.get(statement_name, [])
    position = {name: idx for idx, name in enumerate(order)}
    return sorted(sections, key=lambda s: (position.get(s["section"], 999), s["section"]))


def get_trial_balance(fiscal_year_id: int | None = None, date_from: str | None = None, date_to: str | None = None):
    session = _new_session()
    try:
        ctx = _journal_context(session)
        required = [ctx["jl_gl_fk"], ctx["jl_debit"], ctx["jl_credit"], ctx["jl_batch_fk"], ctx["ga_id"], ctx["ga_name"]]
        if not all(required):
            raise ReportingServiceError("Journal/GL schema is missing expected columns for Trial Balance.")

        filters = []
        if ctx["jb_status"]:
            filters.append(f"COALESCE(jb.{ctx['jb_status']}, 'posted') NOT IN ('draft', 'cancelled')")
        if ctx["jb_fy"]:
            filters.append(f"(:fiscal_year_id IS NULL OR jb.{ctx['jb_fy']} = CAST(:fiscal_year_id AS BIGINT))")
        if ctx["jb_date"] and date_from:
            filters.append(f"jb.{ctx['jb_date']} >= CAST(:date_from AS DATE)")
        if ctx["jb_date"] and date_to:
            filters.append(f"jb.{ctx['jb_date']} <= CAST(:date_to AS DATE)")
        where_clause = ("WHERE " + " AND ".join(filters)) if filters else ""
        code_select = f"ga.{ctx['ga_code']} AS account_code," if ctx["ga_code"] else "'' AS account_code,"

        sql = f"""
            SELECT
                ga.{ctx['ga_id']} AS gl_account_id,
                {code_select}
                ga.{ctx['ga_name']} AS account_name,
                COALESCE(SUM(jl.{ctx['jl_debit']}), 0) AS total_debit,
                COALESCE(SUM(jl.{ctx['jl_credit']}), 0) AS total_credit,
                COALESCE(SUM(jl.{ctx['jl_debit']}) - SUM(jl.{ctx['jl_credit']}), 0) AS net_balance
            FROM journal_lines jl
            JOIN gl_accounts ga ON ga.{ctx['ga_id']} = jl.{ctx['jl_gl_fk']}
            JOIN journal_batches jb ON jb.id = jl.{ctx['jl_batch_fk']}
            {where_clause}
            GROUP BY ga.{ctx['ga_id']}, ga.{ctx['ga_name']}{", ga." + ctx["ga_code"] if ctx["ga_code"] else ""}
            ORDER BY ga.{ctx['ga_code'] if ctx['ga_code'] else ctx['ga_name']}, ga.{ctx['ga_name']}
        """
        rows = session.execute(text(sql), {"fiscal_year_id": fiscal_year_id, "date_from": date_from, "date_to": date_to}).mappings().all()
        rows = [dict(r) for r in rows]
        totals = {
            "total_debit": sum(Decimal(str(r.get("total_debit", 0) or 0)) for r in rows),
            "total_credit": sum(Decimal(str(r.get("total_credit", 0) or 0)) for r in rows),
        }
        return {"rows": rows, "totals": totals}
    finally:
        session.close()


def _gfs_statement_rows(statement_name: str, fiscal_year_id: int | None = None, date_from: str | None = None, date_to: str | None = None):
    session = _new_session()
    try:
        ctx = _journal_context(session)
        required = [ctx["jl_gl_fk"], ctx["jl_debit"], ctx["jl_credit"], ctx["jl_batch_fk"], ctx["ga_id"], ctx["gm_ga_fk"], ctx["gm_gfs_fk"], ctx["g_id"], ctx["g_name"], ctx["g_type"], ctx["g_stmt"]]
        if not all(required):
            raise ReportingServiceError(f"Schema is missing expected columns for {statement_name}.")

        filters = [f"g.{ctx['g_stmt']} = :statement_name"]
        if ctx["jb_status"]:
            filters.append(f"COALESCE(jb.{ctx['jb_status']}, 'posted') NOT IN ('draft', 'cancelled')")
        if ctx["jb_fy"]:
            filters.append(f"(:fiscal_year_id IS NULL OR jb.{ctx['jb_fy']} = CAST(:fiscal_year_id AS BIGINT))")
        if ctx["jb_date"] and date_from:
            filters.append(f"jb.{ctx['jb_date']} >= CAST(:date_from AS DATE)")
        if ctx["jb_date"] and date_to:
            filters.append(f"jb.{ctx['jb_date']} <= CAST(:date_to AS DATE)")

        sql = f"""
            SELECT
                COALESCE(g.{ctx['g_section']}, 'Unclassified') AS statement_section,
                g.{ctx['g_code']} AS gfs_code,
                g.{ctx['g_name']} AS gfs_name,
                COALESCE(g.{ctx['g_note']}, '') AS note_no,
                COALESCE(g.{ctx['g_group']}, '') AS report_group,
                g.{ctx['g_type']} AS gfs_type,
                COALESCE(SUM(jl.{ctx['jl_debit']}), 0) AS total_debit,
                COALESCE(SUM(jl.{ctx['jl_credit']}), 0) AS total_credit,
                CASE
                    WHEN g.{ctx['g_type']} IN ('asset', 'expenditure') THEN COALESCE(SUM(jl.{ctx['jl_debit']}) - SUM(jl.{ctx['jl_credit']}), 0)
                    WHEN g.{ctx['g_type']} IN ('liability', 'equity', 'revenue', 'contra_asset') THEN COALESCE(SUM(jl.{ctx['jl_credit']}) - SUM(jl.{ctx['jl_debit']}), 0)
                    ELSE COALESCE(SUM(jl.{ctx['jl_debit']}) - SUM(jl.{ctx['jl_credit']}), 0)
                END AS statement_amount
            FROM journal_lines jl
            JOIN journal_batches jb ON jb.id = jl.{ctx['jl_batch_fk']}
            JOIN gl_accounts ga ON ga.{ctx['ga_id']} = jl.{ctx['jl_gl_fk']}
            JOIN gl_account_gfs_map gm ON gm.{ctx['gm_ga_fk']} = ga.{ctx['ga_id']}
            JOIN gfs_codes g ON g.{ctx['g_id']} = gm.{ctx['gm_gfs_fk']}
            WHERE {" AND ".join(filters)}
            GROUP BY g.{ctx['g_code']}, g.{ctx['g_name']}, g.{ctx['g_type']}, g.{ctx['g_section']}, g.{ctx['g_note']}, g.{ctx['g_group']}
            ORDER BY COALESCE(g.{ctx['g_section']}, 'Unclassified'), g.{ctx['g_code']}
        """
        rows = session.execute(text(sql), {"statement_name": statement_name, "fiscal_year_id": fiscal_year_id, "date_from": date_from, "date_to": date_to}).mappings().all()
        return [dict(r) for r in rows]
    finally:
        session.close()


def get_statement(statement_name: str, fiscal_year_id: int | None = None, prior_fiscal_year_id: int | None = None, date_from: str | None = None, date_to: str | None = None):
    current_rows = _gfs_statement_rows(statement_name, fiscal_year_id, date_from, date_to)
    prior_rows = _gfs_statement_rows(statement_name, prior_fiscal_year_id, None, None) if prior_fiscal_year_id else []
    prior_index = {r["gfs_code"]: Decimal(str(r.get("statement_amount", 0) or 0)) for r in prior_rows}
    sections: dict[str, dict[str, Any]] = {}
    seen_codes = set()

    for r in current_rows:
        section = r["statement_section"] or "Unclassified"
        current_amt = Decimal(str(r.get("statement_amount", 0) or 0))
        prior_amt = prior_index.get(r["gfs_code"], Decimal("0.00"))
        seen_codes.add(r["gfs_code"])
        sections.setdefault(section, {"section": section, "rows": [], "current_total": Decimal("0.00"), "prior_total": Decimal("0.00")})
        sections[section]["rows"].append({**r, "current_amount": current_amt, "prior_amount": prior_amt})
        sections[section]["current_total"] += current_amt
        sections[section]["prior_total"] += prior_amt

    for r in prior_rows:
        if r["gfs_code"] in seen_codes:
            continue
        section = r["statement_section"] or "Unclassified"
        prior_amt = Decimal(str(r.get("statement_amount", 0) or 0))
        sections.setdefault(section, {"section": section, "rows": [], "current_total": Decimal("0.00"), "prior_total": Decimal("0.00")})
        sections[section]["rows"].append({**r, "current_amount": Decimal("0.00"), "prior_amount": prior_amt})
        sections[section]["prior_total"] += prior_amt

    ordered_sections = _sort_sections(statement_name, list(sections.values()))
    grand_current = sum((s["current_total"] for s in ordered_sections), Decimal("0.00"))
    grand_prior = sum((s["prior_total"] for s in ordered_sections), Decimal("0.00"))
    return {"statement_name": statement_name, "sections": ordered_sections, "grand_current": grand_current, "grand_prior": grand_prior}


def get_changes_in_equity(fiscal_year_id: int | None = None, prior_fiscal_year_id: int | None = None, date_from: str | None = None, date_to: str | None = None):
    current_sofp_rows = _gfs_statement_rows("SOFP", fiscal_year_id, date_from, date_to)
    prior_sofp_rows = _gfs_statement_rows("SOFP", prior_fiscal_year_id, None, None) if prior_fiscal_year_id else []
    current_perf = get_statement("SOFPERF", fiscal_year_id, prior_fiscal_year_id, date_from, date_to)
    current_equity = [r for r in current_sofp_rows if r.get("gfs_type") == "equity"]
    prior_equity_idx = {r["gfs_code"]: Decimal(str(r.get("statement_amount", 0) or 0)) for r in prior_sofp_rows if r.get("gfs_type") == "equity"}

    rows = []
    for r in current_equity:
        rows.append({"gfs_name": r["gfs_name"], "current_amount": Decimal(str(r.get("statement_amount", 0) or 0)), "prior_amount": prior_equity_idx.get(r["gfs_code"], Decimal("0.00"))})

    return {
        "rows": rows,
        "current_surplus_deficit": current_perf["grand_current"],
        "prior_surplus_deficit": current_perf["grand_prior"],
        "current_total_equity": sum((Decimal(str(r["current_amount"])) for r in rows), Decimal("0.00")),
        "prior_total_equity": sum((Decimal(str(r["prior_amount"])) for r in rows), Decimal("0.00")),
    }


def get_cash_flow_statement(date_from: str | None = None, date_to: str | None = None, prior_date_from: str | None = None, prior_date_to: str | None = None):
    def compute_window(df: str | None, dt: str | None):
        session = _new_session()
        try:
            cr_cols = _get_columns(session, "cash_receipts")
            crl_cols = _get_columns(session, "cash_receipt_lines")
            cp_cols = _get_columns(session, "cash_payments")
            cpl_cols = _get_columns(session, "cash_payment_lines")
            cr_status = _pick_first_existing(cr_cols, "status")
            cr_date = _pick_first_existing(cr_cols, "receipt_date", "transaction_date")
            cr_id = _pick_first_existing(cr_cols, "id")
            crl_fk = _pick_first_existing(crl_cols, "cash_receipt_id", "receipt_id")
            crl_amt = _pick_first_existing(crl_cols, "amount")
            cp_status = _pick_first_existing(cp_cols, "status")
            cp_date = _pick_first_existing(cp_cols, "payment_date", "transaction_date")
            cp_id = _pick_first_existing(cp_cols, "id")
            cpl_fk = _pick_first_existing(cpl_cols, "cash_payment_id", "payment_id")
            cpl_amt = _pick_first_existing(cpl_cols, "amount")
            total_receipts = Decimal("0.00")
            total_payments = Decimal("0.00")

            if all([cr_id, crl_fk, crl_amt]):
                filters = []
                if cr_status:
                    filters.append(f"COALESCE(cr.{cr_status}, 'posted') = 'posted'")
                if cr_date and df:
                    filters.append(f"cr.{cr_date} >= CAST(:date_from AS DATE)")
                if cr_date and dt:
                    filters.append(f"cr.{cr_date} <= CAST(:date_to AS DATE)")
                where_clause = ("WHERE " + " AND ".join(filters)) if filters else ""
                sql = f"SELECT COALESCE(SUM(crl.{crl_amt}), 0) FROM cash_receipts cr JOIN cash_receipt_lines crl ON crl.{crl_fk} = cr.{cr_id} {where_clause}"
                total_receipts = Decimal(str(session.execute(text(sql), {"date_from": df, "date_to": dt}).scalar() or 0))

            if all([cp_id, cpl_fk, cpl_amt]):
                filters = []
                if cp_status:
                    filters.append(f"COALESCE(cp.{cp_status}, 'posted') = 'posted'")
                if cp_date and df:
                    filters.append(f"cp.{cp_date} >= CAST(:date_from AS DATE)")
                if cp_date and dt:
                    filters.append(f"cp.{cp_date} <= CAST(:date_to AS DATE)")
                where_clause = ("WHERE " + " AND ".join(filters)) if filters else ""
                sql = f"SELECT COALESCE(SUM(cpl.{cpl_amt}), 0) FROM cash_payments cp JOIN cash_payment_lines cpl ON cpl.{cpl_fk} = cp.{cp_id} {where_clause}"
                total_payments = Decimal(str(session.execute(text(sql), {"date_from": df, "date_to": dt}).scalar() or 0))

            return {"receipts": total_receipts, "payments": total_payments, "net": total_receipts - total_payments}
        finally:
            session.close()

    current = compute_window(date_from, date_to)
    prior = compute_window(prior_date_from, prior_date_to)
    return {
        "current_receipts": current["receipts"],
        "current_payments": current["payments"],
        "current_net": current["net"],
        "prior_receipts": prior["receipts"],
        "prior_payments": prior["payments"],
        "prior_net": prior["net"],
    }


def get_notes_index():
    session = _new_session()
    try:
        cols = _get_columns(session, "gfs_codes")
        if not cols:
            return []
        code_col = _pick_first_existing(cols, "code")
        name_col = _pick_first_existing(cols, "name")
        stmt_col = _pick_first_existing(cols, "statement_name")
        note_col = _pick_first_existing(cols, "note_no")
        section_col = _pick_first_existing(cols, "statement_section")
        sql = f"""
            SELECT
                COALESCE({note_col}, '') AS note_no,
                {code_col} AS code,
                {name_col} AS name,
                COALESCE({stmt_col}, '') AS statement_name,
                COALESCE({section_col}, '') AS statement_section
            FROM gfs_codes
            WHERE COALESCE({note_col}, '') <> ''
            ORDER BY {note_col}, {code_col}
        """
        rows = session.execute(text(sql)).mappings().all()
        return [dict(r) for r in rows]
    finally:
        session.close()


def get_note_schedule(note_no: str, fiscal_year_id: int | None = None, prior_fiscal_year_id: int | None = None):
    session = _new_session()
    try:
        ctx = _journal_context(session)
        required = [
            ctx["jl_gl_fk"],
            ctx["jl_debit"],
            ctx["jl_credit"],
            ctx["jl_batch_fk"],
            ctx["ga_id"],
            ctx["ga_name"],
            ctx["gm_ga_fk"],
            ctx["gm_gfs_fk"],
            ctx["g_id"],
            ctx["g_name"],
            ctx["g_note"],
            ctx["g_type"],
        ]
        if not all(required):
            raise ReportingServiceError("Schema is missing expected columns for note schedules.")

        def fetch_rows(fy_id):
            filters = [f"COALESCE(g.{ctx['g_note']}, '') = :note_no"]
            if ctx["jb_status"]:
                filters.append(f"COALESCE(jb.{ctx['jb_status']}, 'posted') NOT IN ('draft', 'cancelled')")
            if ctx["jb_fy"]:
                filters.append(f"(:fiscal_year_id IS NULL OR jb.{ctx['jb_fy']} = CAST(:fiscal_year_id AS BIGINT))")

            ga_code_expr = f"COALESCE(ga.{ctx['ga_code']}, '')" if ctx["ga_code"] else "''"

            sql = f"""
                SELECT
                    {ga_code_expr} AS gl_code,
                    ga.{ctx['ga_name']} AS gl_name,
                    g.{ctx['g_code']} AS gfs_code,
                    g.{ctx['g_name']} AS gfs_name,
                    g.{ctx['g_type']} AS gfs_type,
                    COALESCE(SUM(jl.{ctx['jl_debit']}), 0) AS total_debit,
                    COALESCE(SUM(jl.{ctx['jl_credit']}), 0) AS total_credit,
                    CASE
                        WHEN g.{ctx['g_type']} IN ('asset', 'expenditure') THEN COALESCE(SUM(jl.{ctx['jl_debit']}) - SUM(jl.{ctx['jl_credit']}), 0)
                        WHEN g.{ctx['g_type']} IN ('liability', 'equity', 'revenue', 'contra_asset') THEN COALESCE(SUM(jl.{ctx['jl_credit']}) - SUM(jl.{ctx['jl_debit']}), 0)
                        ELSE COALESCE(SUM(jl.{ctx['jl_debit']}) - SUM(jl.{ctx['jl_credit']}), 0)
                    END AS note_amount
                FROM journal_lines jl
                JOIN journal_batches jb ON jb.id = jl.{ctx['jl_batch_fk']}
                JOIN gl_accounts ga ON ga.{ctx['ga_id']} = jl.{ctx['jl_gl_fk']}
                JOIN gl_account_gfs_map gm ON gm.{ctx['gm_ga_fk']} = ga.{ctx['ga_id']}
                JOIN gfs_codes g ON g.{ctx['g_id']} = gm.{ctx['gm_gfs_fk']}
                WHERE {" AND ".join(filters)}
                GROUP BY
                    ga.{ctx['ga_name']},
                    {f'ga.{ctx["ga_code"]},' if ctx["ga_code"] else ''}
                    g.{ctx['g_code']},
                    g.{ctx['g_name']},
                    g.{ctx['g_type']}
                ORDER BY ga.{ctx['ga_name']}
            """
            return [
                dict(r)
                for r in session.execute(
                    text(sql),
                    {"note_no": note_no, "fiscal_year_id": fy_id},
                ).mappings().all()
            ]

        current_rows = fetch_rows(fiscal_year_id)
        prior_rows = fetch_rows(prior_fiscal_year_id) if prior_fiscal_year_id else []

        prior_index = {
            (r["gl_code"], r["gfs_code"]): Decimal(str(r.get("note_amount", 0) or 0))
            for r in prior_rows
        }

        rows = []
        seen = set()

        for r in current_rows:
            key = (r["gl_code"], r["gfs_code"])
            seen.add(key)
            rows.append({
                **r,
                "current_amount": Decimal(str(r.get("note_amount", 0) or 0)),
                "prior_amount": prior_index.get(key, Decimal("0.00")),
            })

        for r in prior_rows:
            key = (r["gl_code"], r["gfs_code"])
            if key in seen:
                continue
            rows.append({
                **r,
                "current_amount": Decimal("0.00"),
                "prior_amount": Decimal(str(r.get("note_amount", 0) or 0)),
            })

        return {
            "note_no": note_no,
            "rows": rows,
            "current_total": sum((Decimal(str(r["current_amount"])) for r in rows), Decimal("0.00")),
            "prior_total": sum((Decimal(str(r["prior_amount"])) for r in rows), Decimal("0.00")),
        }
    finally:
        session.close()


def export_statement_to_excel(title: str, result: dict):
    Workbook, Font = _load_openpyxl()

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)

    row_no = 3
    if "sections" in result:
        ws.cell(row=row_no, column=1, value="Code")
        ws.cell(row=row_no, column=2, value="Description")
        ws.cell(row=row_no, column=3, value="Current Year")
        ws.cell(row=row_no, column=4, value="Previous Year")
        for c in range(1, 5):
            ws.cell(row=row_no, column=c).font = Font(bold=True)
        row_no += 1

        for section in result["sections"]:
            ws.cell(row=row_no, column=1, value=section["section"]).font = Font(bold=True)
            row_no += 1
            for item in section["rows"]:
                ws.cell(row=row_no, column=1, value=item.get("gfs_code"))
                ws.cell(row=row_no, column=2, value=item.get("gfs_name"))
                ws.cell(row=row_no, column=3, value=float(item.get("current_amount", 0) or 0))
                ws.cell(row=row_no, column=4, value=float(item.get("prior_amount", 0) or 0))
                row_no += 1
            ws.cell(row=row_no, column=2, value="Section Total").font = Font(bold=True)
            ws.cell(row=row_no, column=3, value=float(section.get("current_total", 0) or 0)).font = Font(bold=True)
            ws.cell(row=row_no, column=4, value=float(section.get("prior_total", 0) or 0)).font = Font(bold=True)
            row_no += 2

    elif "rows" in result:
        headers = list(result["rows"][0].keys()) if result["rows"] else []
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=row_no, column=idx, value=h).font = Font(bold=True)
        row_no += 1
        for item in result["rows"]:
            for idx, h in enumerate(headers, start=1):
                val = item[h]
                if isinstance(val, Decimal):
                    val = float(val)
                ws.cell(row=row_no, column=idx, value=val)
            row_no += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out